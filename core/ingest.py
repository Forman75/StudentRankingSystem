from __future__ import annotations
import csv
from io import BytesIO, StringIO
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import load_workbook
# =========================

# Excel: читаем лист как матрицу, разворачиваем merged cells
# =========================
def _sheet_to_matrix_with_merged(wb_bytes: bytes, sheet_name: str, max_rows: Optional[int] = None) -> List[List[Any]]:
    wb = load_workbook(BytesIO(wb_bytes), read_only=False, data_only=True)
    ws = wb[sheet_name]
    merged_map = {}
    for r in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = r.bounds
        top_val = ws.cell(min_row, min_col).value
        for rr in range(min_row, max_row + 1):
            for cc in range(min_col, max_col + 1):
                merged_map[(rr, cc)] = top_val

    rows = []
    max_r = ws.max_row
    max_c = ws.max_column
    if max_rows is not None:
        max_r = min(max_r, max_rows)

    for r in range(1, max_r + 1):
        row_vals = []
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if (r, c) in merged_map and (v is None or str(v).strip() == ""):
                v = merged_map[(r, c)]
            row_vals.append(v)
        rows.append(row_vals)

    return rows
# =========================

# CSV: устойчивое чтение из bytes (формы/выгрузки)
# =========================
def _decode_sample(data: bytes, enc: str, limit: int = 65536) -> str:
    # Декодирует кусок текста для sniff delimiter / отладки
    try:
        return data[:limit].decode(enc, errors="replace")
    except Exception:
        return data[:limit].decode("utf-8", errors="replace")


def _guess_delimiter(sample_text: str) -> str:
    # разделитель Forms: ',' (en-US) или ';' (ru locales), иногда табы, csv.Sniffer
    try:
        dialect = csv.Sniffer().sniff(sample_text, delimiters=";,\t|")
        if dialect.delimiter:
            return dialect.delimiter
    except Exception:
        pass

    # fallback по количеству в первых строках
    candidates = [";", ",", "\t", "|"]
    lines = [ln for ln in sample_text.splitlines() if ln.strip()][:20]
    if not lines:
        return ","

    scores = {}
    for d in candidates:
        # среднее количество разделителей на строку
        cnts = [ln.count(d) for ln in lines]
        scores[d] = sum(cnts) / max(1, len(cnts))

    # выбираем лучший, но если все 0 - пусть будет ','
    best = max(scores.items(), key=lambda x: x[1])[0]
    return best if scores.get(best, 0) > 0 else ","


def _read_csv_bytes(data: bytes) -> pd.DataFrame:
    # читаем CSV БЕЗ header чтобы первая строка (заголовки формы) попадала в df_raw как обычная строка
    encodings = ["utf-8-sig", "utf-8", "cp1251"]
    last_err: Exception | None = None

    for enc in encodings:
        try:
            sample = _decode_sample(data, enc)
            delim = _guess_delimiter(sample)

            # читаем как матрицу
            df = pd.read_csv(
                BytesIO(data),
                header=None,
                sep=delim,
                engine="python",
                encoding=enc,
                skip_blank_lines=True,
            )

            # если вдруг прочитался в 1 колонку
            if df.shape[1] == 1:
                for d2 in [";", ",", "\t", "|"]:
                    if d2 == delim:
                        continue
                    df2 = pd.read_csv(
                        BytesIO(data),
                        header=None,
                        sep=d2,
                        engine="python",
                        encoding=enc,
                        skip_blank_lines=True,
                    )
                    if df2.shape[1] > 1:
                        df = df2
                        break

            return df

        except Exception as e:
            last_err = e
            continue

    # Декодируем как текст с заменой и читаем
    sample = _decode_sample(data, "utf-8")
    delim = _guess_delimiter(sample)
    try:
        df = pd.read_csv(
            StringIO(sample),
            header=None,
            sep=delim,
            engine="python",
            skip_blank_lines=True,
        )
        return df
    except Exception as e:
        # если всё совсем плохо - поднимем исходную ошибку
        raise last_err or e
# =========================

# Main: uploads -> tables
# =========================
def load_tables_from_uploads(uploads) -> List[Dict[str, Any]]:
    """
    Возвращает список таблиц в формате:
      {
        "source_name": <имя файла>,
        "sheet_name": <лист или 'CSV'>,
        "df_raw": DataFrame  (первая колонка - _origin_row),
      }

    В df_raw:
      - CSV читается как "матрица" без заголовков (header=None)
      - Excel читается как матрица (включая merged cells)
      - _origin_row проставляется как "номер строки в исходнике" (начиная с 1)
    """
    tables: List[Dict[str, Any]] = []

    for up in uploads:
        name = up.name
        data = up.getvalue()

        if name.lower().endswith(".csv"):
            df = _read_csv_bytes(data)

            # унификация: добавляем колонку с номером строки в исходнике (начиная с 1)
            df.insert(0, "_origin_row", range(1, len(df) + 1))
            tables.append({"source_name": name, "sheet_name": "CSV", "df_raw": df})
            continue

        # Excel
        xls = pd.ExcelFile(BytesIO(data))
        for sheet in xls.sheet_names:
            try:
                matrix = _sheet_to_matrix_with_merged(data, sheet_name=sheet, max_rows=None)
                df_raw = pd.DataFrame(matrix)
            except Exception:
                # fallback
                df_raw = pd.read_excel(BytesIO(data), sheet_name=sheet, header=None)

            df_raw.insert(0, "_origin_row", range(1, len(df_raw) + 1))
            tables.append({"source_name": name, "sheet_name": sheet, "df_raw": df_raw})

    return tables
